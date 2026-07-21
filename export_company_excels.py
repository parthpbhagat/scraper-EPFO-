from __future__ import annotations

import argparse
import calendar
import datetime as dt
import json
import os
import re
from copy import copy
from pathlib import Path
from typing import Any

import mysql.connector
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment

from epfo_scraper import MYSQL_DEFAULTS, read_company_names


DEFAULT_TEMPLATE = r"C:\Users\BAPS\Downloads\RELIANCE INDUSTRIES LIMITED (1).xlsx"
DEFAULT_OUTPUT_DIR = "outputs/company_epfo_excels"
MONTH_YEAR = "July_2026"
DEFAULT_CIN_OVERRIDES = "company_cin_overrides.json"


def mysql_config() -> dict[str, Any]:
    return {
        "host": os.getenv("MYSQL_HOST", MYSQL_DEFAULTS["MYSQL_HOST"]),
        "port": int(os.getenv("MYSQL_PORT", MYSQL_DEFAULTS["MYSQL_PORT"])),
        "user": os.getenv("MYSQL_USER", MYSQL_DEFAULTS["MYSQL_USER"]),
        "password": os.getenv("MYSQL_PASSWORD", MYSQL_DEFAULTS["MYSQL_PASSWORD"]),
        "database": os.getenv("MYSQL_DATABASE", MYSQL_DEFAULTS["MYSQL_DATABASE"]),
        "charset": "utf8mb4",
        "use_unicode": True,
    }


def clean_label(value: str | None) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip().strip(":").strip()


def clean_filename_part(value: str, max_len: int = 100) -> str:
    value = re.sub(r"[<>:\"/\\|?*\x00-\x1f]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip(" .")
    return (value[:max_len].strip(" .") or "Company")


def amount_to_number(value: str | int | float | None) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    digits = re.sub(r"[^0-9.-]", "", str(value))
    if not digits:
        return 0
    return int(float(digits))


def parse_credit_date(value: str | None) -> dt.date | None:
    if not value:
        return None
    value = value.strip()
    for fmt in ("%d-%b-%Y", "%d-%B-%Y", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(value.title(), fmt).date()
        except ValueError:
            continue
    return None


def parse_wage_month(value: str | None) -> tuple[int, int] | None:
    if not value:
        return None
    text = value.strip().replace("_", "-").replace(" ", "-")
    match = re.match(r"^([A-Za-z]{3,9})-?(\d{2,4})$", text)
    if not match:
        return None
    month_text, year_text = match.groups()
    month_lookup = {name[:3].upper(): idx for idx, name in enumerate(calendar.month_abbr) if name}
    month = month_lookup.get(month_text[:3].upper())
    if not month:
        return None
    year = int(year_text)
    if year < 100:
        year += 2000
    return year, month


def display_date(value: dt.date | None) -> str:
    return value.strftime("%d/%m/%Y") if value else "-"


def due_date_for_wage_month(wage_month: str | None) -> dt.date | None:
    parsed = parse_wage_month(wage_month)
    if not parsed:
        return None
    year, month = parsed
    if month == 12:
        return dt.date(year + 1, 1, 15)
    return dt.date(year, month + 1, 15)


def delay_text(first_credit: dt.date | None, due_date: dt.date | None) -> str:
    if not first_credit or not due_date:
        return "-"
    days = max(0, (first_credit - due_date).days)
    return f"{days} Day"


def wage_sort_key(wage_month: str | None) -> tuple[int, int]:
    parsed = parse_wage_month(wage_month)
    if not parsed:
        return (0, 0)
    return parsed


def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    source_height = ws.row_dimensions[source_row].height
    if source_height is not None:
        ws.row_dimensions[target_row].height = source_height
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def capture_row_style(ws, source_row: int, max_col: int) -> dict[str, Any]:
    return {
        "height": ws.row_dimensions[source_row].height,
        "cells": [
            {
                "style": copy(ws.cell(source_row, col)._style),
                "font": copy(ws.cell(source_row, col).font),
                "fill": copy(ws.cell(source_row, col).fill),
                "border": copy(ws.cell(source_row, col).border),
                "alignment": copy(ws.cell(source_row, col).alignment),
                "number_format": ws.cell(source_row, col).number_format,
                "protection": copy(ws.cell(source_row, col).protection),
            }
            for col in range(1, max_col + 1)
        ],
    }


def apply_row_style(ws, target_row: int, style: dict[str, Any]) -> None:
    if style.get("height") is not None:
        ws.row_dimensions[target_row].height = style["height"]
    for col, cell_style in enumerate(style["cells"], start=1):
        cell = ws.cell(target_row, col)
        cell._style = copy(cell_style["style"])
        cell.font = copy(cell_style["font"])
        cell.fill = copy(cell_style["fill"])
        cell.border = copy(cell_style["border"])
        cell.alignment = copy(cell_style["alignment"])
        cell.number_format = cell_style["number_format"]
        cell.protection = copy(cell_style["protection"])


def clear_sheet_after_header(ws, header_row: int = 3, max_col: int | None = None) -> None:
    max_col = max_col or ws.max_column
    note_ranges = list(ws.merged_cells.ranges)
    for merged_range in note_ranges:
        min_row, _, max_row, _ = merged_range.bounds
        if min_row > header_row or max_row > header_row:
            ws.unmerge_cells(str(merged_range))
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)


def append_note(ws, row: int, note: str, max_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row, 1)
    cell.value = note
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 36


def extract_field_map(records: list[dict[str, Any]]) -> dict[str, str]:
    fields: dict[str, str] = {}
    for record in records:
        field_name = clean_label(record.get("field_name"))
        field_value = record.get("field_value")
        if field_name and field_value not in (None, ""):
            fields[field_name] = str(field_value).strip()

        row_json = record.get("row_json")
        if not row_json:
            continue
        try:
            row = json.loads(row_json)
        except json.JSONDecodeError:
            continue

        label = clean_label(row.get("B."))
        value = row.get("C.")
        if label and value not in (None, ""):
            fields[label] = str(value).strip()

        for first, second in (("column_1", "column_2"), ("column_3", "column_4")):
            label = clean_label(row.get(first))
            value = row.get(second)
            if label and value not in (None, ""):
                fields[label] = str(value).strip()
    return fields


def trust_status(exemption_status: str | None) -> str:
    text = (exemption_status or "").upper()
    if "UNEXEMPTED" in text:
        return "NON PRIVATE TRUST"
    if "EXEMPTED" in text:
        return "PRIVATE TRUST"
    return "NON PRIVATE TRUST"


def meaningful_code(value: str | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.upper() in {"NA", "N/A", "NULL", "-"}:
        return None
    return text


def aggregate_payments(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        est_id = row.get("est_id") or ""
        wage_month = row.get("wage_month") or "-"
        key = (est_id, wage_month)
        target = grouped.setdefault(
            key,
            {
                "est_id": est_id,
                "wage_month": wage_month,
                "amount": 0,
                "no_of_employee": 0,
                "first_credit_date": None,
            },
        )
        target["amount"] += amount_to_number(row.get("amount"))
        target["no_of_employee"] += amount_to_number(row.get("no_of_employee"))
        credit_date = parse_credit_date(row.get("date_of_credit"))
        if credit_date and (target["first_credit_date"] is None or credit_date < target["first_credit_date"]):
            target["first_credit_date"] = credit_date

    return sorted(
        grouped.values(),
        key=lambda item: (wage_sort_key(item["wage_month"]), item["first_credit_date"] or dt.date.min),
        reverse=True,
    )


def latest_run_id(cursor) -> int:
    cursor.execute("SELECT MAX(id) AS run_id FROM scrape_runs")
    row = cursor.fetchone()
    return int(row["run_id"] or 0)


def load_cin_overrides(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    return {str(key).strip().upper(): str(value).strip() for key, value in data.items() if str(value).strip()}


def fetch_company_payload(cursor, run_id: int, company_name: str) -> dict[str, Any]:
    if run_id:
        cursor.execute(
            """
            SELECT *
            FROM company_queries
            WHERE run_id = %s AND company_name = %s
            ORDER BY id DESC
            LIMIT 1
            """,
            (run_id, company_name),
        )
    else:
        cursor.execute(
            """
            SELECT *
            FROM company_queries
            WHERE company_name = %s
            ORDER BY (status = 'completed') DESC, run_id DESC, id DESC
            LIMIT 1
            """,
            (company_name,),
        )
    query = cursor.fetchone()
    if not query:
        return {"company_name": company_name, "query": None, "establishments": []}

    data_run_id = int(query["run_id"])

    cursor.execute(
        """
        SELECT e.*
        FROM company_query_establishments cqe
        JOIN establishments e ON e.est_id = cqe.est_id
        WHERE cqe.query_id = %s
        ORDER BY e.est_id
        """,
        (query["id"],),
    )
    establishments = cursor.fetchall()

    payload_establishments = []
    for establishment in establishments:
        est_id = establishment["est_id"]
        cursor.execute(
            """
            SELECT *
            FROM establishment_section_data
            WHERE est_id = %s AND (run_id = %s OR query_id = %s)
            ORDER BY section_name, table_no, row_no, id
            """,
            (est_id, data_run_id, query["id"]),
        )
        sections = cursor.fetchall()
        field_map = extract_field_map(sections)

        cursor.execute(
            """
            SELECT *
            FROM payment_details
            WHERE est_id = %s
            ORDER BY id
            """,
            (est_id,),
        )
        payments = cursor.fetchall()
        payload_establishments.append(
            {
                "establishment": establishment,
                "fields": field_map,
                "payments": aggregate_payments(payments),
            }
        )

    return {"company_name": company_name, "query": query, "establishments": payload_establishments}


def write_epf_sheet(ws, payload: dict[str, Any], cin_overrides: dict[str, str]) -> str:
    company_name = payload["company_name"]
    ws["A1"] = company_name.upper()
    data_row_style = capture_row_style(ws, 4, 10)
    note_row_style = capture_row_style(ws, ws.max_row if ws.max_row >= 4 else 3, 10)
    clear_sheet_after_header(ws, header_row=3, max_col=10)
    ws.auto_filter.ref = "A3:J3"
    ws.freeze_panes = "A4"

    output_rows: list[list[Any]] = []
    cin = cin_overrides.get(company_name.upper(), "NA")
    for item in payload["establishments"]:
        establishment = item["establishment"]
        fields = item["fields"]
        payments = item["payments"]
        cin = meaningful_code(fields.get("CIN Code")) or cin
        state_or_office = fields.get("State") or establishment.get("office_name") or "-"
        working_status = fields.get("Working Status") or "-"
        trust = trust_status(fields.get("Exemption Status"))

        for payment in payments:
            due_date = due_date_for_wage_month(payment["wage_month"])
            first_credit = payment["first_credit_date"]
            output_rows.append(
                [
                    payment["est_id"],
                    trust,
                    working_status,
                    state_or_office,
                    payment["wage_month"],
                    display_date(due_date),
                    display_date(first_credit),
                    delay_text(first_credit, due_date),
                    payment["amount"],
                    payment["no_of_employee"],
                ]
            )

    if output_rows:
        for row_index, row_values in enumerate(output_rows, start=4):
            apply_row_style(ws, row_index, data_row_style)
            for col_index, value in enumerate(row_values, start=1):
                ws.cell(row_index, col_index).value = value
        last_row = 3 + len(output_rows)
        ws.auto_filter.ref = f"A3:J{last_row}"
    else:
        apply_row_style(ws, 4, data_row_style)
        append_note(ws, 4, "No EPFO payment rows found in the project database for this company.", 10)

    note_start = ws.max_row + 2
    apply_row_style(ws, note_start, note_row_style)
    append_note(
        ws,
        note_start,
        "* EPFO data exported from the local provident_fund database. Please double-check the source before consuming.",
        10,
    )
    apply_row_style(ws, note_start + 2, note_row_style)
    append_note(
        ws,
        note_start + 2,
        "# The data that we get from the EPFO website is often not up to date at the source itself.",
        10,
    )
    return cin or "NA"


def write_gst_sheet(ws, company_name: str) -> None:
    ws["A1"] = company_name.upper()
    data_row_style = capture_row_style(ws, 4, 8)
    clear_sheet_after_header(ws, header_row=3, max_col=8)
    ws.auto_filter.ref = "A3:H3"
    ws.freeze_panes = "A4"
    apply_row_style(ws, 4, data_row_style)
    append_note(ws, 4, "GST data is not available in this EPFO scraper project database.", 8)


def save_company_workbook(
    template: Path,
    output_dir: Path,
    payload: dict[str, Any],
    month_year: str,
    cin_overrides: dict[str, str],
) -> Path:
    wb = load_workbook(template)
    epf = wb["EPF"] if "EPF" in wb.sheetnames else wb.active
    cin = write_epf_sheet(epf, payload, cin_overrides)
    if "GST" in wb.sheetnames:
        write_gst_sheet(wb["GST"], payload["company_name"])

    company_part = clean_filename_part(payload["company_name"], 100)
    cin_part = clean_filename_part(cin, 30)
    filename = f"{company_part}_EPFO_{cin_part}_{month_year}.xlsx"
    output_path = output_dir / filename
    output_dir.mkdir(parents=True, exist_ok=True)
    for old_path in output_dir.glob(f"{company_part}_EPFO_*_{month_year}.xlsx"):
        if old_path.name.startswith("~$") or old_path == output_path:
            continue
        old_path.unlink()
    wb.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export EPFO database data into sample-style company Excel files.")
    parser.add_argument("--company-file", default="compony.txt", help="Company list file.")
    parser.add_argument("--template", default=DEFAULT_TEMPLATE, help="Sample Excel workbook path.")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="Folder for generated Excel files.")
    parser.add_argument("--month-year", default=MONTH_YEAR, help="Month_Year suffix for filenames.")
    parser.add_argument(
        "--run-id",
        type=int,
        default=0,
        help="Scrape run id. 0 uses latest completed query per company when available.",
    )
    parser.add_argument("--cin-overrides", default=DEFAULT_CIN_OVERRIDES, help="Optional JSON file of company-to-CIN overrides.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    template = Path(args.template)
    if not template.exists():
        raise FileNotFoundError(f"Template workbook not found: {template}")

    companies = read_company_names(Path(args.company_file))
    if not companies:
        raise RuntimeError(f"No companies found in {args.company_file}")

    conn = mysql.connector.connect(**mysql_config())
    try:
        cursor = conn.cursor(dictionary=True)
        selected_run_id = args.run_id
        latest_id = latest_run_id(cursor)
        if not latest_id:
            raise RuntimeError("No scrape run found in database.")
        cin_overrides = load_cin_overrides(Path(args.cin_overrides))

        output_dir = Path(args.output_dir)
        created: list[Path] = []
        for company in companies:
            payload = fetch_company_payload(cursor, selected_run_id, company)
            created.append(save_company_workbook(template, output_dir, payload, args.month_year, cin_overrides))

        print(f"Run ID used: {selected_run_id or 'latest completed per company'}")
        print(f"Created {len(created)} workbook(s):")
        for path in created:
            print(path.resolve())
        cursor.close()
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
